from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
RAIZ = Path(__file__).parent
WORKBOOK_PATH = RAIZ / 'workbook.xlsx'

workbook = Workbook()
worksheet = workbook.active

# listar abas da planilha
# print(workbook.sheetnames)
# criar abas na planilha 
sheet_name = 'aba 2'
workbook.create_sheet(sheet_name)

# ir para uma aba da planilha
# worksheet = workbook[sheet_name]
# remover abas 
# workbook.remove[workbook[sheet_name]]

# criando cabecalhos 
worksheet.cell(1,1,'Nome') # type: ignore
worksheet.cell(1,2,'Idade') # type: ignore
worksheet.cell(1,3,'Nota') # type: ignore

estudantes = [
    ['joao',14,5.5],
    ['Maria',13,9.7],
    ['Gabriel',28,10.5],
    ['Barbara',22,2]
]

# inserindo dados 
'''for i,student_row in enumerate(estudantes,start=2):
    for j,student_column in enumerate(student_row, start=1):
        worksheet.cell(i, j, student_column) # type: ignore
'''
# inserindo dados de outra forma
for student in estudantes:
    worksheet.append(student) # type: ignore
workbook.save(WORKBOOK_PATH) # type: ignore


# lendo/ alterando uma planilha
JOGOS_RATE_PATH = RAIZ / 'jogos_rate.xlsx'
# carregando o xlsx
jogos_rate = load_workbook(JOGOS_RATE_PATH)
jogos_sheet = jogos_rate['jogos']
# for na tabela inteira
'''for i,row in enumerate(jogos_sheet.iter_rows(), start= 1):
    print(f' linha {i}' )
    for col in row:
        print(col.value)
'''        
# ver dados de uma coluna e linha especifica
print(jogos_sheet['A17'].value)
# alterar dados de uma coluna e linha expecifica 
jogos_sheet['A17'].value = 'pokemon violettt'
print(jogos_sheet['A17'].value)